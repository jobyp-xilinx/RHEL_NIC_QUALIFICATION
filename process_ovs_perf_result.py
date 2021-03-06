import argparse
import csv
import glob
import io
import os
import re
import sys
import tarfile
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing import image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import colors

if sys.version_info[0] == 3:
    raw_input = input

DPDK_L3_PVP_PNGS = ['test_p2v2p_all_l3_ref.png',
                    'test_p2v2p_all_l3.png',
                    'test_p2v2p_1000000_l3.png',
                    'test_p2v2p_100000_l3.png',
                    'test_p2v2p_10000_l3.png',
                    'test_p2v2p_1000_l3.png',
                    'test_p2v2p_10_l3.png',]

DPDK_L2_PVP_PNGS = ['test_p2v2p_all_l2_ref.png',
                    'test_p2v2p_all_l2.png',
                    'test_p2v2p_1000000_l2.png',
                    'test_p2v2p_100000_l2.png',
                    'test_p2v2p_10000_l2.png',
                    'test_p2v2p_1000_l2.png',
                    'test_p2v2p_10_l2.png',]

KERNEL_L3_PVP_PNGS = ['test_p2v2p_all_l3_ref.png',
                      'test_p2v2p_all_l3.png',
                      'test_p2v2p_1000000_l3.png',
                      'test_p2v2p_100000_l3.png',
                      'test_p2v2p_10000_l3.png',
                      'test_p2v2p_1000_l3.png',
                      'test_p2v2p_10_l3.png',]

KERNEL_L2_PVP_PNGS = ['test_p2v2p_all_l2_ref.png',
                      'test_p2v2p_all_l2.png',
                      'test_p2v2p_1000000_l2.png',
                      'test_p2v2p_100000_l2.png',
                      'test_p2v2p_10000_l2.png',
                      'test_p2v2p_1000_l2.png',
                      'test_p2v2p_10_l2.png',]

TC_L3_PVP_PNGS = ['test_p2v2p_all_l3_ref.png',
                  'test_p2v2p_all_l3.png',
                  'test_p2v2p_1000000_l3.png',
                  'test_p2v2p_100000_l3.png',
                  'test_p2v2p_10000_l3.png',
                  'test_p2v2p_1000_l3.png',
                  'test_p2v2p_10_l3.png',]

TC_L2_PVP_PNGS = ['test_p2v2p_all_l2_ref.png',
                  'test_p2v2p_all_l2.png',
                  'test_p2v2p_1000000_l2.png',
                  'test_p2v2p_100000_l2.png',
                  'test_p2v2p_10000_l2.png',
                  'test_p2v2p_1000_l2.png',
                  'test_p2v2p_10_l2.png',]


class ResultsSheet(object):
    def __init__(self, args):
        """
        Constructor
        :param args: arguments parse object from command line output
        """
        self._args = args
        self.result_tar_file = args.result_tar_file
        self.sheet_name = self.result_tar_file.split(".")[0]
        self.worksheet = None
        output_file = self._args.output
        if os.path.exists(output_file):
            self.workbook = load_workbook(output_file)
            self.save_file = output_file
        elif os.path.exists(os.getcwd() + os.sep + output_file):
            cur_path = os.getcwd() + os.sep + output_file
            self.workbook = load_workbook(cur_path)
            self.save_file = cur_path
        else:
            self.workbook = Workbook()
            self.workbook.save(output_file)
            self.save_file = output_file
            self.worksheet = self.workbook.active
        if None == self.worksheet:
            self.worksheet = self.workbook.create_sheet(self.sheet_name)
        self.row = 1
        self.workbook.save(self.save_file)

    def close_workbook(self):
        """
        Close the workbook
        :return: None
        """
        self.workbook.save(self.save_file)
        self.workbook.close()

    def process_pvp_results(self):
        """
        Process Eelcos pvp results
        :return: None
        """
        tar = tarfile.open(self.result_tar_file,"r:*")
        if 'dpdk' in self.result_tar_file:
            if "l2" in self.result_tar_file:
                if self.write_pvp_worksheet(tar, f'{self.sheet_name}/test_results_l2.csv',
                                            self.worksheet, DPDK_L2_PVP_PNGS):
                    self.worksheet.title = self.sheet_name + ' (FAIL)'
                else:
                    self.worksheet.title = self.sheet_name + ' (PASS)'
            elif "l3" in self.result_tar_file:
                if self.write_pvp_worksheet(tar, f'{self.sheet_name}/test_results_l3.csv',
                                            self.worksheet, DPDK_L3_PVP_PNGS):
                    self.worksheet.title = self.sheet_name + ' (FAIL)'
                else:
                    self.worksheet.title = self.sheet_name + ' (PASS)'
            else:
                pass
        # find the kernel result file and process it
        elif 'kernel' in self.result_tar_file:
            if "l2" in self.result_tar_file:
                if self.write_pvp_worksheet(tar, f'{self.sheet_name}/test_results_l2.csv',
                                            self.worksheet, KERNEL_L2_PVP_PNGS):
                    self.worksheet.title = self.sheet_name + ' (FAIL)'
                else:
                    self.worksheet.title = self.sheet_name + ' (PASS)'
            elif "l3" in self.result_tar_file:
                if self.write_pvp_worksheet(tar, f'{self.sheet_name}/test_results_l3.csv',
                                            self.worksheet, KERNEL_L3_PVP_PNGS):
                    self.worksheet.title = self.sheet_name + ' (FAIL)'
                else:
                    self.worksheet.title = self.sheet_name + ' (PASS)'
            else:
                pass
        # find the tc_flower result file and process it
        elif 'tc' in self.result_tar_file:
            #
            # Add TC flower results only if data is available as they are
            # optional.
            #
            if "l2" in self.result_tar_file:
                if self.write_pvp_worksheet(tar, f'{self.sheet_name}/test_results_l2.csv',self.worksheet, TC_L2_PVP_PNGS):
                    self.worksheet.title = self.sheet_name + ' (FAIL)'
                else:
                    self.worksheet.title = self.sheet_name + ' (PASS)'
            elif "l3" in self.result_tar_file:
                if self.write_pvp_worksheet(tar, f'{self.sheet_name}/test_results_l3.csv',self.worksheet, TC_L3_PVP_PNGS):
                    self.worksheet.title = self.sheet_name + ' (FAIL)'
                else:
                    self.worksheet.title = self.sheet_name + ' (PASS)'
            elif "flower" in self.result_tar_file:
                if self.write_tc_throughput_worksheet(tar,f'{self.sheet_name}/test_results_l3.csv'):
                    self.worksheet.title = self.sheet_name + ' (FAIL)'
                else:
                    self.worksheet.title = self.sheet_name + ' (PASS)'
            else:
                pass

    def write_pvp_worksheet(self, tar_file, csv_file, worksheet, png_list):
        """
        Write out the pvp results to the specified worksheet, write pass fail on worksheet name
        :param tar_file: tar file
        :param csv_file: csv file to process inside of tar
        :param worksheet: worksheet to write to
        :param png_list: png list from CONSTANT values
        :return: return boolean if any test failed
        """
        fh = io.StringIO(tar_file.extractfile(csv_file).read().decode('ascii'))
        reader = csv.reader(fh, delimiter=',', quotechar='|')
        test_fail = False
        max_column = 1
        self.row = 1
        for row in reader:
            column = 1
            try:
                if len(row) > 0 and 'cpu' not in row[0]:
                    for i in range(len(row)):
                        # try to convert to int to round.
                        try:
                            entry = int(float(row[i]))
                            if entry <= 0:
                                test_fail = True
                        except ValueError:
                            entry = row[i]
                        temp_cell = worksheet.cell(self.row,column)
                        temp_cell.value = str(entry)
                        column += 1
                    self.row += 1
                    max_column = column if column > max_column else max_column
            except IndexError:
                continue
        self.workbook.save(self.save_file)
        for png in png_list[1:2]:
            tar_file.extractall()
            img = image.Image(self.sheet_name + "/" + png)
            img.anchor = worksheet.cell(self.row,1).coordinate
            worksheet.add_image(img)
        self.workbook.save(self.save_file)

        for col in range(1,max_column):
            worksheet.column_dimensions[get_column_letter(col)].width = 30
        self.row = 1
        self.workbook.save(self.save_file)
        return test_fail

    def write_tc_throughput_worksheet(self, tar_file, csv_file):
        speed_pass_table = {
            10: {64: 2490513, 128: 2400604, 256: 2256274, 512: 2114661,
                 768: 1427664, 1024: 1077586, 1514: 733376},
            25: {64: 9637204, 128: 6508473, 256: 4000375, 512: 3693513,
                 768: 3462503, 1024: 2671544, 1514: 1819386},
            40: {64: 10217438, 128: 9092933, 256: 8657281, 512: 6536093,
                 768: 5280149, 1024: 4024205, 1514: 2619575},
            50: {64: 26067795, 128: 23473241, 256: 18772297, 512: 10476825,
                 768: 7083181, 1024: 5347950, 1514: 3640232},
            100: {64: 26899051, 128: 24033668, 256: 18772297, 512: 13542611,
                  768: 9978782, 1024: 7948678, 1514: 5451593}
        }

        fh = io.StringIO(tar_file.extractfile(csv_file).read().decode('ascii'))
        reader = csv.reader(fh, delimiter=',', quotechar='|')
        nic_speed = None
        packet_sizes = None
        tenK_results = None
        failure = False
        for row in reader:

            if len(row) == 0:
                #
                # Empty line means new test results (we should have only one)
                #
                packet_sizes = None
                tenK_results = None
            elif nic_speed is None:
                #
                # We need the NIC speed
                #
                if row[0] == '\"Physical port':
                    nic_speed = int(float(row[2].split()[1]))
                    if nic_speed not in [10, 25, 40, 50, 100]:
                        raise ValueError(
                            "Unsupported link rate, please report!!")

            elif packet_sizes is None:
                #
                # See if this is the line with the packet size info
                #
                if len(row) >= 2 and row[0] == 'Number of flows':
                    try:
                        packet_sizes = [int(i) for i in row[1:]]
                    except ValueError:
                        packet_sizes = None
            else:
                #
                # Here we are processing results
                #
                if (len(row) == len(packet_sizes) + 1):
                    try:
                        results_values = [int(float(i)) for i in row]
                    except ValueError:
                        if row[0].startswith("cpu_"):
                            #
                            # If line starts with cpu_x it's the embedded
                            # cpu statistics which we should ignore those
                            #
                            continue
                        else:
                            #
                            # Rest are real issues, start over...
                            #
                            packet_sizes = None
                            continue
                    #
                    # Normal result line processing
                    #
                    if results_values[0] == 10000:
                        tenK_results = results_values[1:]
                        break

        if nic_speed is None or packet_sizes is None or tenK_results is None:
            raise ValueError("The TC L3 PVP results is missing data, i.e. "
                             "Link speed, or 10K packet results!!")

        #
        # Write default workbook layout
        #
        for col in range(1,3):
            self.worksheet.column_dimensions[get_column_letter(col)].width = 16

        bold_font = Font(bold=True)

        self.worksheet.cell(1,1).value = "PVP test for 10K TC Flower rules with NIC speed of {} Gbps".format(nic_speed)
        self.worksheet.cell(1,1).font = bold_font

        self.worksheet.cell(2,1).value = "Packet Size"
        self.worksheet.cell(2,1).font = bold_font

        self.worksheet(2,2).value = "pps"
        self.worksheet(2,2).font = bold_font

        self.worksheet(2,3).value = "pass criteria pps"
        self.worksheet(2,3).font = bold_font

        #
        # Write results, and check pass/fail criteria
        #
        red_font = Font(color=colors.RED)
        results = dict(list(zip(packet_sizes, tenK_results)))
        pass_table = speed_pass_table[nic_speed]
        all_packet_sizes = sorted(list(set(packet_sizes + list(pass_table.keys()))))

        for i, pkt_size in enumerate(all_packet_sizes):
            self.worksheet.cell(3,1).value = str(pkt_size)
            try:
                self.worksheet.cell(3,2).value =  str(results[pkt_size])
            except KeyError:
                self.worksheet.cell(3,2).value =  "N/A"
            try:
                if pkt_size not in results or results[pkt_size] < pass_table[pkt_size]:
                    failure = True
                    self.worksheet.cell(i+3,2).value = str(pass_table[pkt_size])
                    self.worksheet.cell(i+3,2).font = red_font
                else:
                    self.worksheet.cell(i+3,2).value = str(pass_table[pkt_size])
            except KeyError:
                self.worksheet.cell(i+3,2).value = "-"

        self.workbook.save(self.save_file)
        return failure

def main():
    mysheet = ResultsSheet(args)
    mysheet.process_pvp_results()
    mysheet.close_workbook()


def yes_no(answer):
    yes = set(['yes', 'y', 'ye', ''])
    no = set(['no', 'n'])

    while True:
        choice = raw_input(answer).lower()
        if choice in yes:
            return True
        elif choice in no:
            return False
        else:
            print("Please respond with 'yes' or 'no'\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-o', '--output', type=str, required=True,
                        help='Output file name')
    parser.add_argument('-s', '--result_tar_file', type=str, required=True,
                        help='result tar file name')
    args = parser.parse_args()
    # if os.path.isfile(args.output):
    #     ans = yes_no("Output file {} already exists. Overwrite?".format(args.output))
    #     if not ans:
    #         sys.exit()
    if os.path.isfile(args.result_tar_file) == False:
        print("Result tar file do not exist. Check your arguments.")
    main()
